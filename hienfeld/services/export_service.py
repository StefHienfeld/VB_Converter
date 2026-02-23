# hienfeld/services/export_service.py
"""
Service for exporting analysis results to various formats.
"""
from typing import Dict, List, Optional, TYPE_CHECKING
from io import BytesIO
import pandas as pd

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from ..config import AppConfig
from ..domain.clause import Clause
from ..domain.cluster import Cluster
from ..domain.analysis import AnalysisAdvice
from ..domain.reference import (
    ReferenceClause,
    ReferenceMatch,
    ComparisonStatus,
    get_comparison_status,
    get_comparison_symbol,
)
from ..logging_config import get_logger

if TYPE_CHECKING:
    from .reference_analysis_service import ReferenceAnalysisService

logger = get_logger('export_service')

# Indicators that mark an action as completed
DONE_INDICATORS = ['ja', 'yes', 'gedaan', 'done', 'x', '✓', '✅', 'afgerond', 'klaar']


class ExportService:
    """
    Handles export of analysis results to Excel and other formats.
    """

    # Cell-level colors for the Advies column only (not full rows)
    ADVICE_COLORS = {
        'VERWIJDEREN': 'FFCDD2',        # Red
        'VERLOPEN': 'FFCDD2',           # Red
        'BEHOUDEN': 'C8E6C9',           # Green
        'HANDMATIG CHECKEN': 'FFE0B2',  # Orange
        'STANDAARDISEREN': 'BBDEFB',    # Blue
        'OPSCHONEN': 'FFF9C4',          # Yellow
        'AANVULLEN': 'FFF9C4',          # Yellow
    }

    # Confidence cell colors
    CONFIDENCE_COLORS = {
        'Hoog': 'C8E6C9',   # Green
        'Midden': 'FFF9C4',  # Yellow
        'Laag': 'FFCDD2',    # Red
    }

    # Header styling
    HEADER_FILL = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
    HEADER_FONT = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    HEADER_ALIGNMENT = Alignment(horizontal='center', vertical='center', wrap_text=True)
    HEADER_BORDER = Border(
        bottom=Side(style='medium', color='2C3E50'),
    )

    # Data cell defaults
    DATA_FONT = Font(name='Calibri', size=10, color='333333')
    DATA_ALIGNMENT = Alignment(vertical='top', wrap_text=False)

    def __init__(self, config: AppConfig):
        """
        Initialize the export service.

        Args:
            config: Application configuration
        """
        self.config = config

    # ------------------------------------------------------------------ #
    #  Excel formatting helpers                                           #
    # ------------------------------------------------------------------ #

    def _apply_excel_formatting(self, ws, sheet_type: str = 'results') -> None:
        """
        Apply professional formatting to an openpyxl worksheet.

        Args:
            ws: openpyxl Worksheet object
            sheet_type: 'results' | 'summary' | 'instructions' | 'detail'
        """
        if ws.max_row is None or ws.max_row < 1:
            return

        # --- 1. Freeze header row + AutoFilter ---
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = ws.dimensions

        # --- 2. Header row styling ---
        for cell in ws[1]:
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = self.HEADER_ALIGNMENT
            cell.border = self.HEADER_BORDER

        # Build column-name-to-index map (1-based)
        col_map = {}
        for cell in ws[1]:
            if cell.value:
                col_map[str(cell.value)] = cell.column

        # --- 3. Column widths ---
        fixed_widths = {'Tekst': 60, 'Reden': 40, 'Artikel': 25}
        for col_cells in ws.columns:
            col_letter = get_column_letter(col_cells[0].column)
            col_idx = col_cells[0].column
            header_name = str(ws.cell(row=1, column=col_idx).value or '')

            if header_name in fixed_widths:
                ws.column_dimensions[col_letter].width = fixed_widths[header_name]
                continue

            # Auto-fit based on content (sample first 100 rows)
            max_length = 0
            for cell in col_cells[:101]:
                try:
                    cell_len = len(str(cell.value or ''))
                    if cell_len > max_length:
                        max_length = cell_len
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max(max_length + 3, 10), 40)

        # --- 4. Data cell styling ---
        # Apply default font + alignment to all data cells
        for row_num in range(2, ws.max_row + 1):
            for col_num in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.font = self.DATA_FONT
                cell.alignment = self.DATA_ALIGNMENT

        # Only Reden gets wrap_text (it's the explanation column)
        reden_idx = col_map.get('Reden')
        if reden_idx:
            wrap_alignment = Alignment(vertical='top', wrap_text=True)
            for row_num in range(2, ws.max_row + 1):
                ws.cell(row=row_num, column=reden_idx).alignment = wrap_alignment


        # Only apply data-row styling for results/detail sheets
        if sheet_type in ('results', 'detail'):
            self._apply_data_row_styling(ws, col_map)

        # --- 5. Status dropdown (data validation) ---
        status_idx = col_map.get('Status')
        if status_idx and ws.max_row > 1:
            dv = DataValidation(
                type='list',
                formula1='"Te beoordelen,Akkoord,Aangepast,Overslaan"',
                allow_blank=True,
            )
            dv.error = 'Kies een geldige status'
            dv.errorTitle = 'Ongeldige status'
            dv.prompt = 'Kies een status'
            dv.promptTitle = 'Status'
            status_letter = get_column_letter(status_idx)
            dv.add(f'{status_letter}2:{status_letter}{ws.max_row}')
            ws.add_data_validation(dv)


    def _apply_data_row_styling(self, ws, col_map: dict) -> None:
        """Apply cell-level color coding to Advies and Vertrouwen columns only."""
        advies_idx = col_map.get('Advies')
        vertrouwen_idx = col_map.get('Vertrouwen')

        for row_num in range(2, ws.max_row + 1):
            # --- Advies cell coloring (only the cell, not the row) ---
            if advies_idx:
                advies_val = str(ws.cell(row=row_num, column=advies_idx).value or '')

                fill_hex = None
                if advies_val.startswith('📋'):
                    fill_hex = 'E1BEE7'  # Purple for custom instructions
                else:
                    for keyword, color in self.ADVICE_COLORS.items():
                        if keyword in advies_val.upper():
                            fill_hex = color
                            break

                if fill_hex:
                    ws.cell(row=row_num, column=advies_idx).fill = PatternFill(
                        start_color=fill_hex, end_color=fill_hex, fill_type='solid'
                    )
                    ws.cell(row=row_num, column=advies_idx).font = Font(
                        name='Calibri', size=10, bold=True, color='333333'
                    )

            # --- Vertrouwen cell coloring (only the cell) ---
            if vertrouwen_idx:
                conf_val = str(ws.cell(row=row_num, column=vertrouwen_idx).value or '')
                conf_hex = self.CONFIDENCE_COLORS.get(conf_val)
                if conf_hex:
                    ws.cell(row=row_num, column=vertrouwen_idx).fill = PatternFill(
                        start_color=conf_hex, end_color=conf_hex, fill_type='solid'
                    )
                    ws.cell(row=row_num, column=vertrouwen_idx).font = Font(
                        name='Calibri', size=10, bold=True, color='333333'
                    )


    def _determine_action_status(self, ref_match: Optional[ReferenceMatch]) -> str:
        """
        Bepaal de actie status op basis van de referentie match.

        Args:
            ref_match: Optional reference match from previous analysis

        Returns:
            Status string: "🆕 Nieuw", "✅ Afgerond", or "🔲 Open"
        """
        if ref_match is None:
            return "🆕 Nieuw"

        ref_status = (ref_match.reference_clause.status or "").strip().lower()

        # Check if the action is marked as completed
        if any(indicator in ref_status for indicator in DONE_INDICATORS):
            return "✅ Afgerond"

        return "🔲 Open"
    
    def build_results_dataframe(
        self,
        clauses: List[Clause],
        clusters: List[Cluster],
        advice_map: Dict[str, AnalysisAdvice],
        include_original_columns: bool = True,
        original_df: Optional[pd.DataFrame] = None,
        hierarchical_results: Optional[List[Dict]] = None,
        reference_matches: Optional[Dict[str, ReferenceMatch]] = None,
        reference_service: Optional['ReferenceAnalysisService'] = None
    ) -> pd.DataFrame:
        """
        Build a DataFrame with analysis results, supporting hierarchical parent/child structure.

        Args:
            clauses: List of analyzed Clause objects (legacy, may be empty if using hierarchical_results)
            clusters: List of Cluster objects
            advice_map: Mapping of cluster_id -> AnalysisAdvice
            include_original_columns: Whether to include original data columns
            original_df: Original DataFrame (for preserving columns)
            hierarchical_results: Optional list of hierarchical result dicts with 'type' ('PARENT', 'CHILD', 'SINGLE')
            reference_matches: Optional mapping of simplified_text -> ReferenceMatch for comparison (deprecated)
            reference_service: Optional reference analysis service for per-clause matching (preferred)

        Returns:
            DataFrame with analysis results
        """
        # Use hierarchical results if available, otherwise fall back to legacy method
        if hierarchical_results:
            return self._build_hierarchical_dataframe(
                hierarchical_results,
                clusters,
                original_df
            )
        
        # Legacy method (backward compatibility)
        logger.info(f"Building results DataFrame from {len(clauses)} clauses (legacy mode)")
        
        # Create cluster lookup for efficiency
        cluster_lookup = {c.id: c for c in clusters}

        # Identify source columns (all except the text column)
        source_cols: List[str] = []
        if include_original_columns and original_df is not None:
            text_col = self._detect_text_column(original_df)
            source_cols = [c for c in original_df.columns if c != text_col]
        
        # Build rows
        rows = []
        for clause in clauses:
            cluster_id = clause.cluster_id or "NVT"
            
            # Get cluster info
            cluster = cluster_lookup.get(cluster_id)
            advice = advice_map.get(cluster_id)
            
            # Get reference match for this clause (if available)
            # Prefer reference_service (per-clause matching with policy_number support)
            ref_match = None
            if reference_service and reference_service.is_loaded:
                # Use reference service directly - matches on text + policy_number
                ref_match = reference_service.find_match(
                    clause.simplified_text,
                    policy_number=clause.source_policy_number
                )
            elif reference_matches:
                # Backward compatibility: use pre-built dict
                simplified = clause.simplified_text.lower().strip() if clause.simplified_text else ""
                ref_match = reference_matches.get(simplified)

            # Build row with standard columns
            # WORKFLOW-FIRST column order: Status, Tekst, Advies, Reden, Artikel, Vertrouwen, Frequentie, Cluster_ID
            row = {
                # Status kolom - initialiseren voor tracking workflow
                'Status': 'Te beoordelen',
                # Analysis output columns (workflow order)
                'Advies': advice.advice_code if advice else '',
                'Reden': advice.reason if advice else '',
                'Artikel': advice.reference_article if advice else '',
                'Vertrouwen': advice.confidence if advice else '',
                # Frequency and cluster info
                'Frequentie': cluster.frequency if cluster else 0,
                'Cluster_ID': cluster_id,
                'Cluster_Naam': cluster.name if cluster else '',
            }

            # Add reference columns (if reference analysis was used)
            if reference_service or reference_matches is not None:
                current_advice = advice.advice_code if advice else ""
                comparison_status = get_comparison_status(current_advice, ref_match)

                row['Actie Status'] = self._determine_action_status(ref_match)
                row['Ref. Frequentie'] = ref_match.reference_clause.frequency if ref_match else ''
                row['Ref. Advies'] = ref_match.reference_clause.advice_code if ref_match else ''
                row['Ref. Status'] = ref_match.reference_clause.status if ref_match else ''
                row['Vergelijking'] = get_comparison_symbol(comparison_status)

            # Add text column
            row['Tekst'] = clause.raw_text
            
            # Add policy number if available
            if clause.source_policy_number:
                row['Polisnummer'] = clause.source_policy_number

            # Add original source columns per policy row (e.g., vervaldatum, product, etc.)
            if include_original_columns and original_df is not None and source_cols:
                orig_idx = self._extract_original_index(clause.id)
                if orig_idx is not None and orig_idx in original_df.index:
                    original_row = original_df.loc[orig_idx]
                    for col in source_cols:
                        try:
                            row[col] = original_row[col]
                        except Exception:
                            row[col] = ""
                else:
                    for col in source_cols:
                        row[col] = ""
            
            rows.append(row)
        
        df = pd.DataFrame(rows)

        # POST-PROCESSING: Group singleton clusters (freq=1) into "Uniek" meta-clusters
        df = self._group_unique_texts(df)

        # Sort by cluster ID for readability
        df = df.sort_values(by='Cluster_ID')

        logger.info(f"Created DataFrame with {len(df)} rows, {len(df.columns)} columns")
        return df

    def _extract_original_index(self, clause_id: str) -> Optional[int]:
        """
        Extract original DataFrame index from Clause ID.

        Supported formats:
        - "row_{idx}"
        - "{policy_number}_{idx}" (policy number may contain underscores; idx is last part)
        """
        if not clause_id:
            return None

        clause_id = str(clause_id)
        if clause_id.startswith('row_'):
            try:
                return int(clause_id.split('_')[1])
            except (ValueError, IndexError):
                return None

        if '_' in clause_id:
            try:
                parts = clause_id.rsplit('_', 1)
                if len(parts) == 2:
                    return int(parts[1])
            except (ValueError, IndexError):
                return None

        return None

    def _group_unique_texts(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Group singleton clusters (Frequentie=1) into "Uniek" meta-clusters.

        Strategy:
        - Clusters with freq >= 2: Keep as-is (real duplicates)
        - Clusters with freq == 1: Group by Advies + Vertrouwen

        Result:
        - Normal clusters: CL-0001, CL-0002, etc.
        - Unique clusters: UNIEK-VERWIJDEREN-Hoog, UNIEK-HANDMATIG CHECKEN-Midden, etc.

        Args:
            df: DataFrame with analysis results

        Returns:
            DataFrame with regrouped unique texts
        """
        if df.empty or 'Frequentie' not in df.columns:
            return df

        # Separate real clusters (freq >= 2) from singletons (freq == 1)
        real_clusters = df[df['Frequentie'] >= 2].copy()
        singletons = df[df['Frequentie'] == 1].copy()

        if singletons.empty:
            logger.info("No singleton clusters to regroup")
            return df

        logger.info(f"Regrouping {len(singletons)} singleton clusters into 'Uniek' meta-clusters")

        # Create unique cluster IDs based on Advies + Vertrouwen
        def create_unique_cluster_id(row):
            advies = row.get('Advies', 'ONBEKEND')
            vertrouwen = row.get('Vertrouwen', 'Onbekend')
            # Clean up advies for ID (remove emojis, special chars)
            advies_clean = advies.replace('✓', '').replace('⚠️', '').replace('🔍', '').strip()
            return f"UNIEK-{advies_clean}-{vertrouwen}"

        def create_unique_cluster_name(row):
            advies = row.get('Advies', 'Onbekend')
            vertrouwen = row.get('Vertrouwen', 'Onbekend')
            return f"Unieke teksten - {advies} ({vertrouwen})"

        # Apply grouping
        singletons['Cluster_ID'] = singletons.apply(create_unique_cluster_id, axis=1)
        singletons['Cluster_Naam'] = singletons.apply(create_unique_cluster_name, axis=1)

        # IMPORTANT: Preserve original frequency BEFORE overwriting
        # This is used by reference_analysis_service to get the real frequency (1)
        # instead of the cluster size (e.g., 625) when loading as reference
        singletons['Orig. Frequentie'] = singletons['Frequentie'].copy()

        # Update Frequentie to reflect group size (per unique cluster)
        unique_cluster_sizes = singletons['Cluster_ID'].value_counts().to_dict()
        singletons['Frequentie'] = singletons['Cluster_ID'].map(unique_cluster_sizes)

        # Combine back together
        result = pd.concat([real_clusters, singletons], ignore_index=True)

        # Log statistics
        unique_groups = singletons['Cluster_ID'].nunique()
        logger.info(f"Created {unique_groups} unique meta-clusters from {len(singletons)} singleton texts")
        logger.info(f"Final: {len(real_clusters)} real clusters + {unique_groups} unique groups = {len(real_clusters) + unique_groups} total cluster groups")

        return result

    def _detect_text_column(self, df: pd.DataFrame) -> Optional[str]:
        """
        Best-effort detection of the free-text column in the original DataFrame.

        1) Try common names (Tekst/Vrije Tekst/etc.)
        2) Fallback: pick the column with the highest median string length in a small sample.
        """
        if df is None or df.empty:
            return None

        # Common names (keep in sync with ingestion/preprocessing expectations)
        text_cols = ['Tekst', 'Vrije Tekst', 'Clausule', 'Text', 'Description']
        for col in text_cols:
            if col in df.columns:
                return col

        # Case-insensitive match
        lower_map = {c.lower(): c for c in df.columns}
        for col in text_cols:
            if col.lower() in lower_map:
                return lower_map[col.lower()]

        # Heuristic fallback (sample for performance)
        sample = df.head(200)
        best_col = None
        best_score = -1.0
        for col in df.columns:
            try:
                series = sample[col]
                # Prefer object-like columns; numeric columns won't win on length anyway
                med = series.astype(str).str.len().median()
                if pd.notna(med) and float(med) > best_score:
                    best_score = float(med)
                    best_col = col
            except Exception:
                continue

        return best_col
    
    def _build_hierarchical_dataframe(
        self,
        hierarchical_results: List[Dict],
        clusters: List[Cluster],
        original_df: Optional[pd.DataFrame] = None
    ) -> pd.DataFrame:
        """
        Build DataFrame from hierarchical results structure.
        
        Args:
            hierarchical_results: List of result dicts with 'type', 'id', 'cluster', 'advice', etc.
            clusters: List of Cluster objects (for lookup)
            original_df: Original DataFrame (for preserving columns)
            
        Returns:
            DataFrame with hierarchical structure
        """
        logger.info(f"Building hierarchical DataFrame from {len(hierarchical_results)} results")
        
        # Create cluster lookup
        cluster_lookup = {c.id: c for c in clusters}
        
        # Identify source columns (all except text column)
        source_cols = []
        if original_df is not None:
            # Find text column name (common names)
            text_cols = ['Tekst', 'Vrije Tekst', 'Clausule', 'Text', 'Description']
            text_col = None
            for col in text_cols:
                if col in original_df.columns:
                    text_col = col
                    break
            
            # All other columns are source columns
            source_cols = [c for c in original_df.columns if c != text_col]
        
        export_rows = []
        
        for item in hierarchical_results:
            item_type = item.get('type', 'SINGLE')
            item_id = item.get('id', 'UNKNOWN')
            cluster = item.get('cluster')
            advice = item.get('advice')
            
            # Base row data
            row = {
                'Type': item_type,  # Add Type column for filtering
                'Cluster_ID': item_id,
                'Advies': advice.advice_code if advice else '',
                'Vertrouwen': advice.confidence if advice else '',
                'Reden': advice.reason if advice else '',
                'Artikel': advice.reference_article if advice else '',
            }
            
            if item_type in ['PARENT', 'SINGLE']:
                # Parent/Single rows: include full cluster info and original data
                if cluster:
                    row['Cluster_Naam'] = cluster.name
                    row['Frequentie'] = cluster.frequency
                    row['Tekst'] = cluster.original_text
                    
                    # Get original row data from DataFrame
                    if original_df is not None:
                        # Find original row by extracting index from clause ID
                        # Clause IDs are formatted as "row_{idx}" or "{policy_number}_{idx}"
                        leader_clause = cluster.leader_clause
                        orig_idx = None
                        
                        # Extract index from clause ID
                        clause_id = leader_clause.id
                        if clause_id.startswith('row_'):
                            try:
                                orig_idx = int(clause_id.split('_')[1])
                            except (ValueError, IndexError):
                                pass
                        elif '_' in clause_id:
                            # Format: {policy_number}_{idx}
                            try:
                                parts = clause_id.rsplit('_', 1)
                                if len(parts) == 2:
                                    orig_idx = int(parts[1])
                            except (ValueError, IndexError):
                                pass
                        
                        # Fallback: try to find by matching text
                        if orig_idx is None:
                            # Find text column
                            text_cols = ['Tekst', 'Vrije Tekst', 'Clausule', 'Text', 'Description']
                            text_col = None
                            for col in text_cols:
                                if col in original_df.columns:
                                    text_col = col
                                    break
                            
                            if text_col:
                                for idx, orig_text in original_df[text_col].items():
                                    if str(orig_text).strip() == cluster.original_text.strip():
                                        orig_idx = idx
                                        break
                        
                        # Get original row data
                        if orig_idx is not None:
                            try:
                                # Use loc to get row by index (works with both numeric and named indices)
                                if orig_idx in original_df.index:
                                    original_row = original_df.loc[orig_idx]
                                    # Add ALL source columns
                                    for col in source_cols:
                                        row[col] = original_row[col]
                                else:
                                    # Index not found, fill with empty
                                    for col in source_cols:
                                        row[col] = ""
                            except (KeyError, IndexError):
                                # Index error, fill with empty
                                for col in source_cols:
                                    row[col] = ""
                        else:
                            # Could not find index, fill with empty
                            for col in source_cols:
                                row[col] = ""
                    else:
                        # No original_df, fill source columns with empty
                        for col in source_cols:
                            row[col] = ""
                    
                    # Add clean text proposal
                    clean_text = item.get('clean_text_proposal', '')
                    row['Nieuwe_Systeem_Tekst'] = clean_text
                    
                    # PARENT rows: show summary of child advices
                    if item_type == 'PARENT':
                        children = item.get('children', [])
                        if children:
                            # Summarize child advice codes
                            child_advice_counts = {}
                            for child in children:
                                child_adv = child.get('advice')
                                if child_adv:
                                    code = child_adv.advice_code
                                    child_advice_counts[code] = child_advice_counts.get(code, 0) + 1
                            
                            # Build summary string
                            summary_parts = [f"{count}x {code}" for code, count in child_advice_counts.items()]
                            summary_str = ", ".join(summary_parts) if summary_parts else "geen onderdelen"
                            
                            row['Advies'] = '⚠️ GESPLITST'
                            row['Reden'] = f"Gesplitst in {len(children)} onderdelen: {summary_str}"
                        else:
                            row['Advies'] = '⚠️ ZIE ONDERSTAANDE DELEN'
                            row['Reden'] = "Bevat meerdere onderdelen. Zie details hieronder."
                else:
                    # Fallback if no cluster
                    row['Cluster_Naam'] = ''
                    row['Frequentie'] = 0
                    row['Tekst'] = ''
                    row['Nieuwe_Systeem_Tekst'] = ''
                    for col in source_cols:
                        row[col] = ""
                
            elif item_type == 'CHILD':
                # Child rows: indent text, no source data
                row['Cluster_Naam'] = ''
                row['Frequentie'] = 0
                row['Tekst'] = f"    ↳ {item.get('text', '')}"  # Indentatie
                row['Nieuwe_Systeem_Tekst'] = ''  # Empty for children
                
                # Fill source columns with empty
                for col in source_cols:
                    row[col] = ""
            
            export_rows.append(row)
        
        df = pd.DataFrame(export_rows)
        
        # Sort by Cluster_ID (will group parent and children together)
        df = df.sort_values(by='Cluster_ID')
        
        logger.info(f"Created hierarchical DataFrame with {len(df)} rows, {len(df.columns)} columns")
        return df
    
    def build_cluster_summary(
        self,
        clusters: List[Cluster],
        advice_map: Dict[str, AnalysisAdvice]
    ) -> pd.DataFrame:
        """
        Build a summary DataFrame with one row per cluster.
        
        Args:
            clusters: List of Cluster objects
            advice_map: Mapping of cluster_id -> AnalysisAdvice
            
        Returns:
            Summary DataFrame
        """
        rows = []
        
        for cluster in clusters:
            advice = advice_map.get(cluster.id)
            
            row = {
                'Cluster_ID': cluster.id,
                'Cluster_Naam': cluster.name,
                'Frequentie': cluster.frequency,
                'Advies': advice.advice_code if advice else '',
                'Vertrouwen': advice.confidence if advice else '',
                'Reden': advice.reason if advice else '',
                'Artikel': advice.reference_article if advice else '',
                'Voorbeeld_Tekst': cluster.original_text[:200] + '...' if len(cluster.original_text) > 200 else cluster.original_text
            }
            rows.append(row)
        
        df = pd.DataFrame(rows)
        df = df.sort_values(by='Cluster_ID')
        
        return df
    
    def _sanitize_for_excel(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Sanitize DataFrame for Excel export by removing illegal characters.
        
        openpyxl doesn't allow certain control characters and some Unicode ranges.
        This function cleans all string columns.
        """
        import re
        
        # Pattern for illegal Excel characters (control chars except tab, newline, carriage return)
        # See: https://docs.microsoft.com/en-us/openspecs/office_standards/ms-xlsx/
        illegal_chars = re.compile(r'[\x00-\x08\x0b-\x0c\x0e-\x1f]')
        
        def clean_string(val):
            if isinstance(val, str):
                # Remove illegal control characters
                val = illegal_chars.sub('', val)
                # Replace problematic Unicode chars with ASCII equivalents
                val = val.replace('–', '-')  # en-dash
                val = val.replace('—', '-')  # em-dash
                val = val.replace(''', "'")  # smart quote
                val = val.replace(''', "'")  # smart quote
                val = val.replace('"', '"')  # smart quote
                val = val.replace('"', '"')  # smart quote
                val = val.replace('…', '...')  # ellipsis
            return val
        
        # Apply to all object (string) columns
        df = df.copy()
        for col in df.select_dtypes(include=['object']).columns:
            df[col] = df[col].apply(clean_string)
        
        return df
    
    def to_excel_bytes(
        self,
        df: pd.DataFrame,
        include_summary: bool = False,
        clusters: Optional[List[Cluster]] = None,
        advice_map: Optional[Dict[str, AnalysisAdvice]] = None,
        gone_texts: Optional[List[ReferenceClause]] = None
    ) -> bytes:
        """
        Export DataFrame to Excel bytes.

        Args:
            df: Main results DataFrame
            include_summary: Whether to include a summary sheet
            clusters: Clusters for summary (required if include_summary=True)
            advice_map: Advice map for summary (required if include_summary=True)
            gone_texts: List of reference clauses not found in current data (verdwenen teksten)

        Returns:
            Excel file as bytes
        """
        output = BytesIO()
        
        # Sanitize input DataFrame
        df = self._sanitize_for_excel(df)

        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Split long texts (>800 characters) into separate sheet
            max_text_length = self.config.analysis_rules.max_text_length  # Default: 800

            # Identify long text rows (check if Reden mentions "te lang" or actual text length)
            # FIXED: Check both text length AND reason to catch all long texts
            long_text_mask = (
                (df['Tekst'].str.len() > max_text_length) |
                (df['Reden'].str.contains('te lang', case=False, na=False))
            )

            # Split into two DataFrames
            long_texts_df = df[long_text_mask].copy()
            normal_df = df[~long_text_mask].copy()

            # Write normal results to main sheet
            normal_df.to_excel(writer, sheet_name='Analyseresultaten', index=False)
            self._apply_excel_formatting(writer.sheets['Analyseresultaten'], sheet_type='results')
            logger.info(f"Analyseresultaten sheet: {len(normal_df)} rows")

            # Write long texts to separate sheet (if any)
            if not long_texts_df.empty:
                long_texts_df.to_excel(writer, sheet_name='Lange teksten', index=False)
                self._apply_excel_formatting(writer.sheets['Lange teksten'], sheet_type='results')
                logger.info(f"Lange teksten sheet: {len(long_texts_df)} rows (>{max_text_length} characters)")
            else:
                logger.info("No long texts to separate")

            # Optional summary sheet
            if include_summary and clusters and advice_map:
                summary_df = self.build_cluster_summary(clusters, advice_map)
                summary_df = self._sanitize_for_excel(summary_df)  # Fix: sanitize summary too
                summary_df.to_excel(writer, sheet_name='Cluster Samenvatting', index=False)
                self._apply_excel_formatting(writer.sheets['Cluster Samenvatting'], sheet_type='summary')

            # Verdwenen Teksten sheet (texts in reference but not in current)
            if gone_texts:
                gone_df = self._build_gone_texts_dataframe(gone_texts)
                gone_df = self._sanitize_for_excel(gone_df)  # Fix: sanitize gone texts too
                gone_df.to_excel(writer, sheet_name='Verdwenen Teksten', index=False)
                self._apply_excel_formatting(writer.sheets['Verdwenen Teksten'], sheet_type='summary')
                logger.info(f"Verdwenen Teksten sheet: {len(gone_texts)} rows")

            # UNIEKE_Detail sheet - detailed view of all singleton texts
            # Use normal_df (not df) to exclude long texts that are already in "Lange teksten"
            unique_detail_df = self._build_unique_detail_dataframe(normal_df)
            if not unique_detail_df.empty:
                unique_detail_df = self._sanitize_for_excel(unique_detail_df)
                unique_detail_df.to_excel(writer, sheet_name='Unieke_Detail', index=False)
                self._apply_excel_formatting(writer.sheets['Unieke_Detail'], sheet_type='detail')
                logger.info(f"Unieke_Detail sheet: {len(unique_detail_df)} rows")

            # Instructies sheet - help for colleagues
            instructions_df = self._build_instructions_dataframe()
            instructions_df.to_excel(writer, sheet_name='Instructies', index=False)
            self._apply_excel_formatting(writer.sheets['Instructies'], sheet_type='instructions')
            logger.info("Added Instructies sheet")

        logger.info("Generated Excel file")
        return output.getvalue()

    def _build_gone_texts_dataframe(
        self,
        gone_texts: List[ReferenceClause]
    ) -> pd.DataFrame:
        """
        Build DataFrame for "Verdwenen Teksten" sheet.

        These are texts that existed in the reference analysis but
        are not present in the current data (possibly due to policy
        cancellations, mutations, or text changes).

        Args:
            gone_texts: List of unmatched reference clauses

        Returns:
            DataFrame with gone text information
        """
        rows = []
        for clause in gone_texts:
            rows.append({
                'Status': '🗑️ Verdwenen',
                'Cluster_Naam': clause.cluster_name,
                'Ref. Frequentie': clause.frequency,
                'Ref. Advies': clause.advice_code,
                'Ref. Status': clause.status,
                'Ref. Vertrouwen': clause.confidence,
                'Tekst': clause.text[:500] + '...' if len(clause.text) > 500 else clause.text,
                'Opmerking': 'Niet meer aanwezig in huidige data',
            })

        return pd.DataFrame(rows)

    def _build_unique_detail_dataframe(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Build DataFrame for "Unieke_Detail" sheet.

        This sheet provides a detailed view of all singleton (frequency=1) texts
        that were grouped into UNIEK meta-clusters. This allows colleagues to
        see the individual texts within each UNIEK group.

        Args:
            df: Main results DataFrame (with UNIEK clusters already created)

        Returns:
            DataFrame with detailed unique texts, or empty DataFrame if no unique texts
        """
        if df.empty or 'Cluster_ID' not in df.columns:
            return pd.DataFrame()

        # Filter only UNIEK clusters
        unique_mask = df['Cluster_ID'].str.startswith('UNIEK-', na=False)
        unique_df = df[unique_mask].copy()

        if unique_df.empty:
            return pd.DataFrame()

        # Select and reorder columns for clarity
        detail_columns = [
            'Cluster_ID',  # UNIEK group identifier
            'Status',
            'Tekst',
            'Advies',
            'Reden',
            'Artikel',
            'Vertrouwen',
        ]

        # Filter to existing columns
        available_cols = [c for c in detail_columns if c in unique_df.columns]
        result = unique_df[available_cols].copy()

        # Sort by cluster group then by text for readability
        result = result.sort_values(by=['Cluster_ID', 'Tekst'] if 'Tekst' in result.columns else ['Cluster_ID'])

        logger.info(f"Built Unieke_Detail with {len(result)} rows from {unique_df['Cluster_ID'].nunique()} UNIEK groups")
        return result

    def _build_instructions_dataframe(self) -> pd.DataFrame:
        """
        Build DataFrame for "Instructies" sheet.

        Provides guidance to colleagues on how to use the analysis results.

        Returns:
            DataFrame with instructions and column explanations
        """
        instructions = [
            {'Onderwerp': 'Doel', 'Uitleg': 'Dit rapport bevat de analyse van polisclausules, gegroepeerd in clusters.'},
            {'Onderwerp': 'Status kolom', 'Uitleg': 'Gebruik deze kolom om de voortgang bij te houden. Opties: Te beoordelen, In behandeling, Akkoord, Afgewezen'},
            {'Onderwerp': '', 'Uitleg': ''},
            {'Onderwerp': 'KOLOM UITLEG', 'Uitleg': ''},
            {'Onderwerp': 'Status', 'Uitleg': 'Trackingstatus voor handmatige beoordeling'},
            {'Onderwerp': 'Tekst', 'Uitleg': 'De originele clausuletekst uit de polis'},
            {'Onderwerp': 'Advies', 'Uitleg': 'Aanbevolen actie (VERWIJDEREN, BEHOUDEN, HANDMATIG CHECKEN, etc.)'},
            {'Onderwerp': 'Reden', 'Uitleg': 'Uitleg waarom dit advies wordt gegeven'},
            {'Onderwerp': 'Artikel', 'Uitleg': 'Referentie naar het artikel in de voorwaarden (indien van toepassing)'},
            {'Onderwerp': 'Vertrouwen', 'Uitleg': 'Betrouwbaarheid van het advies: Hoog, Midden, of Laag'},
            {'Onderwerp': 'Frequentie', 'Uitleg': 'Hoe vaak deze tekst voorkomt in de dataset'},
            {'Onderwerp': 'Cluster_ID', 'Uitleg': 'Technische identificatie van de clustergroep'},
            {'Onderwerp': 'Cluster_Naam', 'Uitleg': 'Beschrijvende naam van de cluster'},
            {'Onderwerp': '', 'Uitleg': ''},
            {'Onderwerp': 'ADVIES CODES', 'Uitleg': ''},
            {'Onderwerp': 'VERWIJDEREN', 'Uitleg': 'Tekst kan verwijderd worden - al gedekt door voorwaarden'},
            {'Onderwerp': 'BEHOUDEN (CLAUSULE)', 'Uitleg': 'Tekst moet behouden worden - is een specifieke clausule'},
            {'Onderwerp': 'BEHOUDEN (MAATWERK)', 'Uitleg': 'Tekst moet behouden worden - is maatwerk voor deze polis'},
            {'Onderwerp': 'HANDMATIG CHECKEN', 'Uitleg': 'Automatische analyse onzeker - handmatige beoordeling nodig'},
            {'Onderwerp': 'STANDAARDISEREN', 'Uitleg': 'Tekst komt vaak voor - maak er een standaard clausulecode van'},
            {'Onderwerp': '', 'Uitleg': ''},
            {'Onderwerp': 'SHEETS', 'Uitleg': ''},
            {'Onderwerp': 'Analyseresultaten', 'Uitleg': 'Hoofdresultaten - alle normale clausules'},
            {'Onderwerp': 'Lange teksten', 'Uitleg': 'Clausules die te lang zijn voor automatische analyse'},
            {'Onderwerp': 'Unieke_Detail', 'Uitleg': 'Gedetailleerde weergave van unieke (1x voorkomende) teksten'},
            {'Onderwerp': 'Cluster Samenvatting', 'Uitleg': 'Overzicht per cluster (optioneel)'},
            {'Onderwerp': 'Verdwenen Teksten', 'Uitleg': 'Teksten die in referentie stonden maar nu ontbreken (indien referentie gebruikt)'},
        ]
        return pd.DataFrame(instructions)

    def to_csv_bytes(self, df: pd.DataFrame, delimiter: str = ';') -> bytes:
        """
        Export DataFrame to CSV bytes.
        
        Args:
            df: DataFrame to export
            delimiter: CSV delimiter (default: ';' for Dutch Excel)
            
        Returns:
            CSV file as bytes
        """
        output = BytesIO()
        df.to_csv(output, sep=delimiter, index=False, encoding='utf-8-sig')
        return output.getvalue()
    
    def get_statistics_summary(
        self,
        clauses: List[Clause],
        clusters: List[Cluster],
        advice_map: Dict[str, AnalysisAdvice]
    ) -> dict:
        """
        Generate statistics summary for display.
        
        Args:
            clauses: All clauses
            clusters: All clusters
            advice_map: All advice
            
        Returns:
            Dictionary with statistics
        """
        total_rows = len(clauses)
        unique_clusters = len(clusters)
        
        # Count by advice type
        advice_counts = {}
        category_counts = {}
        found_in_conditions = 0
        
        for advice in advice_map.values():
            code = advice.advice_code
            advice_counts[code] = advice_counts.get(code, 0) + 1
            
            # Track categories
            cat = advice.category or "UNKNOWN"
            category_counts[cat] = category_counts.get(cat, 0) + 1
            
            # Track items found in conditions (KRITIEKE METRIC!)
            if cat and 'VOORWAARDEN' in cat:
                found_in_conditions += 1
        
        # Reduction percentage
        reduction = int((1 - unique_clusters / total_rows) * 100) if total_rows > 0 else 0
        
        # Multi-clause count
        multi_clause_count = sum(1 for c in clauses if c.is_multi_clause)
        
        return {
            'total_rows': total_rows,
            'unique_clusters': unique_clusters,
            'reduction_percentage': reduction,
            'multi_clause_count': multi_clause_count,
            'advice_distribution': advice_counts,
            'category_distribution': category_counts,
            'found_in_conditions': found_in_conditions,
            'avg_cluster_size': total_rows / unique_clusters if unique_clusters > 0 else 0
        }
    
    def format_column_selection(
        self,
        df: pd.DataFrame,
        text_col: str
    ) -> pd.DataFrame:
        """
        Select and order columns for final output.

        Uses WORKFLOW-FIRST ordering: Status, Tekst, Advies, Reden, Artikel, Vertrouwen, Frequentie, then technical columns.

        Args:
            df: Full DataFrame
            text_col: Name of the original text column

        Returns:
            DataFrame with selected columns in workflow order
        """
        # Define desired column order (WORKFLOW-FIRST)
        priority_cols = [
            'Status',           # 1. Tracking
            'Tekst',            # 2. What is it?
            'Advies',           # 3. What to do?
            'Reden',            # 4. Why?
            'Artikel',          # 5. Reference
            'Vertrouwen',       # 6. How sure?
            'Frequentie',       # 7. How often?
            'Cluster_ID',       # 8. Technical ID
            'Cluster_Naam',     # 9. Technical name
        ]

        # Handle text column alias
        if 'Tekst' not in df.columns and text_col in df.columns:
            # Rename text_col to Tekst for consistency
            df = df.rename(columns={text_col: 'Tekst'})

        # Filter to existing columns
        existing_cols = [c for c in priority_cols if c in df.columns]

        # Add any remaining columns
        other_cols = [c for c in df.columns if c not in existing_cols]

        return df[existing_cols + other_cols]

